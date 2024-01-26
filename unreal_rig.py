'''
script to convert a sig rig to a rig that can be used in Unreal Engine
we need to keep the body and head blendshape data so they can be animated
but we want to remove the skinning info so we can re-bind the geo to the UE rig
this contains just the animation skeleton (no def rig)
we want to preserve the blendshape data so we can't just delete history
but if we don't delete history AND we remove the skinCluster connection, the geo might pop back to a different state
instead we will keep the actor's proportions (apply_fitting_data)
and then apply_fitting_data to the incoming source skeleton so the proportions match
then we can transfer the weights
if we restore the characterPose at that point, the geo should change just fine
'''
 
import pymel.all as pm
import maya.cmds as cmds
import maya.mel as mel
import os
import csv
from importlib import reload
import json
import random as rand
import time
import constants
import lib.p4util as p4util
import openpyxl
import rigging.characterBuilder.bodyCsvManager as bodyCSV
import character.assetbuild.vcassets as vcassets
import rigging.animSkeleton.character_pose as cpose
reload(cpose)
import animation.utils.anim_utils as anim_utils
reload(anim_utils)
import shotgun.sg_anim_utils as sg_anim_utils
import animation.rigging.rig_utils as rig_utils
reload(rig_utils)
import assetlibrary.asset_utils as asset_utils
 
DATABUILD_ROOT = str(constants.ART_ROOT)
 
# define some key paths
ACC_MAP   = constants.P4_VCEDEV + "graphics/code/character/character_player_roster_accessory_map.inc"
PLYR_DATA = constants.P4_VCEDEV + "Basketball/data/Roster/roster/PlayerData.xlsx"
TEAM_DATA = constants.P4_VCEDEV + "Basketball/data/Roster/roster/TeamData.xlsx"
ACC_DATA  = constants.P4_VCEDEV + "Basketball/data/Roster/roster/AccessoryData.xlsx"
ACC_DIR   = constants.CHARACTER_DIR + "/male/clothing/player_accessories"
UNI_DIR   = constants.P4_DELIVERYART + "characters/common_data/uniform/uniform_variables.csv"
 
# uniform items
SHORTS_GEO_PATH = constants.CHARACTER_DIR + "/male/clothing/uniforms/shortsReg_nike/average/shortsReg_nike.ma"

# JERSEY_GEO_PATH = constants.CHARACTER_DIR + "/male/clothing/uniforms/jerseyReg_nike_u/average/jerseyReg_nike_u.ma"
SHORTS_TXT_PATH = constants.CHARACTER_DIR + "/male/clothing/uniforms/textures/%s/current/core_away/core_away_s_color.tga"
JERSEY_TXT_PATH = constants.CHARACTER_DIR + "/male/clothing/uniforms/textures/%s/current/core_away/core_away_j_color.tga"
SHOE_DIR = constants.CHARACTER_DIR + "/male/clothing/shoes_court"
UNI_PATH = constants.CHARACTER_DIR + "/male/clothing/uniforms"

# adidas_harden1/average/textures/adidas_harden1_editable_2026/2026_shoe.xml
DEF_ROOT = "def_hips"
ANIM_ROOT = "hips"
MALE_RIG = "male_rig"
butColor = [0.5, 0.5, 0.5]
saveColor = [0.5, 0.7, 1]
auditColor = [1, 0.5, 0.2]
black = [0,0,0]
TASK_TAG = "Generate FBX with Outfit"
UPDATE_CL_DESC = "Updating Mocap Rig"
CREATE_CL_DESC = "Creating Mocap Rig"
KICKBACK = "kb"
READY = "rdy"
 
# JERSEY MAPPING
JERSEY_DICT = {
    "tri_neck"              : "jerseyReg_classic_tri",
    "u_neck"                : "jerseyReg_classic_u",
    "v_neck"                : "jerseyReg_classic_v",
    "wishbone"              : "jerseyReg_classic_wishbone",
    "nike_u_neck"           : "jerseyReg_nike_u",
    "nike_v_neck"           : "jerseyReg_nike_v",
    "nike_wishbone_neck"    : "jerseyReg_nike_wishbone",
}
 
SHORTS_DICT = {
    "CLASSIC"   : "shortsReg_classic",
    "NIKE"      : "shortsReg_nike",
    "RETRO"     : "shortsReg_classic_retro",
}
 
SOCKS_DICT = {
    "CREW"          : "sock_crew",
    "CREWSCRUNCH"   : "sock_crew_scrunch",
    "QUARTER"       : "sock_quarter",
    "TALL"          : "sock_tall",
    "TALLSCRUNCH"   : "sock_tall_scrunch",
    "TALLSQUASH"    : "sock_tall_squash",
}
 
DEF_ROOT = "def_hips"
ANIM_ROOT = "hips"
jointMap = [
    'hips',
    'spine',
    'spine1',
    'spine2',
    'leftArm',
    'leftForearm',
    'leftHand',
    'rightArm',
    'rightForearm',
    'rightHand',
    'neck',
    'head',
    'leftUpLeg',
    'leftLeg',
    'leftFoot',
    'leftToeBase',
    'rightUpLeg',
    'rightLeg',
    'rightFoot',
    'rightToeBase',
]
 
SIG_DATA = str(constants.ART_ROOT) + "/characters/signature_data/"
 
RIGS = {
    "male": "male_UE_source.ma",
    "female": "female_UE_source.ma",
    "child": "child_UE_source.ma",
}
 
DEFAULT_CHEST    = constants.CHARACTER_DIR + "/common_data/textures/skin/chest_color_08.tga"
EYES_TXT_DEFAULT = constants.CHARACTER_DIR + "/common_data/_art_assets/images/eye_texture.tga"
FEMALE = 0
MALE = 1
CHILD = 2
GENDER_STR = {
    FEMALE : "FEMALE",
    MALE   : "MALE",
    CHILD  : "CHILD",
}
 
CROWD_SRC_RIG = {
    "FEMALE" : "z:/crowd_female_src.ma",
    "MALE"   : "z:/crowd_male_src.ma",
}
 
# XL SHEET COLUMN #s
HEADBAND = 4
UNDERSHIRT = 8
SOCKS = 13
SHOES = 18
SHORTS = 21
WRIST_L = 27
WRIST_R = 31
ELBOW_L = 35
ELBOW_R = 39
ARM_L = 43
ARM_R = 48
LEG_L = 52
KNEE_L = 56
LEG_R = 61
KNEE_R = 66
ANKLE_L = 70
ANKLE_R = 74
FINGER_L = 78
FINGER_R = 82
ARMTAPE_L = 86
ARMTAPE_R = 90
GOGGLES = 94
FACEMASK = 96
KEY_NAMES = {
    HEADBAND    : "Headband",
    UNDERSHIRT  : "Undershirt",
    SOCKS       : "Socks",
    SHOES       : "Shoes",
    SHORTS      : "Shorts Type",
    WRIST_L     : "Left Wrist",
    WRIST_R     : "Right Wrist",
    ELBOW_L     : "Left Elbow",
    ELBOW_R     : "Right Elbow",
    ARM_L       : "Left Arm",
    ARM_R       : "Right Arm",
    LEG_L       : "Left Leg",
    LEG_R       : "Right Leg",
    KNEE_L      : "Left Knee",
    KNEE_R      : "Right Knee",
    ANKLE_L     : "Left Ankle",
    ANKLE_R     : "Right Ankle",
    FINGER_L    : "Left Finger",
    FINGER_R    : "Right Finger",
    ARMTAPE_L   : "Left Arm Tape",
    ARMTAPE_R   : "Right Arm Tape",
    GOGGLES     : "Goggles",
    FACEMASK    : "Facemask",
}
 
# GUIDE FOR KEEPING GEO

# if the "key" is part of the outfit geo, then we should hang on to all of the geo in the "value"
FEMALE_KEEP_GEO = {
    "female_blazer_opn_geo"         :   ["body_6543_arms_midsleeve", "body_9987_arms_midsleeve"],
    "female_flannel_geo"            :   ["body_6543_arms_midsleeve", "body_9987_arms_midsleeve"],
    "female_hoodie_pullover_geo"    :   ["body_6543_arms_longsleeve", "body_9987_arms_longsleeve"],
    "female_jersey_basketball_nike_u_male_geo"    :   ["female_dickie_tshirt_ss_geo", "body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "female_jersey_basketball_nike_v_male_geo"    :   ["female_dickie_tshirt_ss_geo", "body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "female_jersey_basketball_nike_wishbone_male_geo"    :   ["female_dickie_tshirt_ss_geo", "body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "female_sweater_geo"            :   ["body_6543_arms_longsleeve", "body_9987_arms_longsleeve"],
    "female_tshirt_ls_geo"          :   ["body_6543_arms_longsleeve", "body_9987_arms_longsleeve"],
    "female_tshirt_ss_finals_geo"   :   ["body_6543_arms_longsleeve", "body_9987_arms_longsleeve"],
    "female_tshirt_ss_geo"          :   ["body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "female_business_geo"           :   ["body_6543_legs_pants", "body_9987_legs_pants"],
    "female_business_geo"           :   ["body_6543_legs_pants", "body_9987_legs_pants"],
    "female_cargo_geo"              :   ["body_6543_legs_pants", "body_9987_legs_pants"],
    "female_jeans_geo"              :   ["body_6543_legs_pants", "body_9987_legs_pants"],
}

MALE_KEEP_GEO = {
    "basketball_shorts_nike_geo"    :   ["body_6543_legs_shorts", "body_9987_legs_shorts"],
    "blazer_opn_geo"                :   ["bdshirt_opn_geo", "body_6543_arms_longsleeve", "body_9987_arms_longsleeve"],
    "business_geo"                  :   ["body_6543_legs_pants", "body_9987_legs_pants"],
    "flannel_geo"                   :   ["body_6543_arms_midsleeve", "body_9987_arms_midsleeve"],
    "hoodie_pullover_geo"           :   ["body_6543_arms_longsleeve", "body_9987_arms_longsleeve"],
    "jacket_warmup_historic_geo"    :   ["body_6543_arms_longsleeve", "body_9987_arms_longsleeve"],
    "jeans_geo"                     :   ["body_6543_legs_pants", "body_9987_legs_pants"],
    "jersey_basketball_nike_u_geo"  :   ["dickie_tshirt_ss_geo", "body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "jersey_basketball_nike_v_geo"  :   ["dickie_tshirt_ss_geo", "body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "jersey_basketball_nike_wishbone_geo"  :   ["dickie_tshirt_ss_geo", "body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "jogger_geo"                    :   ["body_6543_legs_pants", "body_9987_legs_pants"],
    "tshirt_ls_geo"                 :   ["body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "tshirt_ss_finals_geo"          :   ["body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
    "tshirt_ss_geo"                 :   ["body_6543_arms_shortsleeve", "body_9987_arms_shortsleeve"],
}

KEEP_GEO = {}
KEEP_GEO["FEMALE"] = FEMALE_KEEP_GEO
KEEP_GEO["MALE"]   = MALE_KEEP_GEO
 
# make sure user has the USD plugin loaded
if not pm.pluginInfo('mayaUsdPlugin.mll', query=1, loaded=1):
    pm.loadPlugin('mayaUsdPlugin.mll', quiet=1)

# make sure user has the fbx plugin loaded
if not pm.pluginInfo('fbxmaya.mll', query=1, loaded=1):
    pm.loadPlugin('fbxmaya.mll', quiet=1)
CROWD_FILE = constants.CHARACTER_DIR + "/crowd/nba_crowd.vcif"
EXPORT_RIG_PATH = r"z:/Crowd/Rigs"
 
#########################
# CROWD GEO STUFF
#########################
 
def copy_head_weights():
    # quick method to run in Maya
    # select all head geo you want to copy the src blend_head weighting to
    # then run the method to have it update the weighting
    print("copy_head_weights:")
    src_sc = None
    src_shape = None
    blend_head = pm.ls("*blend_head*", type="mesh")
    print("blend_head: ", blend_head)
    if blend_head:
        blend_head = blend_head[0]
        print("blend_head: ", blend_head)
        sc = pm.listConnections(blend_head, type="skinCluster")
        print("connections to %s: %s" % (blend_head, sc))

        if sc:
            src_sc = sc[0]
            print("\t BAC: found a skinCluster %s on %s" % (src_sc, blend_head))

    if src_sc:
        print("src_sc: ", src_sc)
        for head in pm.ls(sl=1):
            print("working on head: ", head)
            shapes = pm.listRelatives(head, c=1)
            print("shapes: ", shapes)
            if shapes:
                shape = shapes[0]
                print("shape: ", shape)
                dst_sc = pm.listConnections(shape, type="skinCluster")
                if dst_sc:
                    dst_sc = dst_sc[0]
                    pm.copySkinWeights(ss=src_sc, ds=dst_sc, sa='closestPoint', noMirror=True)

def copy_body_weights():
    # quick method to run in Maya
    # select all body geo you want to copy the src baseBody weighting to
    # then run the method to have it update the weighting
    print("copy_body_weights:")
    src_sc = None
    src_shape = None
    base_body = pm.ls("*baseBody*", type="mesh")
    print("base_body: ", base_body)
    if base_body:
        base_body = base_body[0]
        print("base_body: ", base_body)
        sc = pm.listConnections(base_body, type="skinCluster")
        print("connections to %s: %s" % (base_body, sc))
        if sc:
            src_sc = sc[0]
            print("\t BAC: found a skinCluster %s on %s" % (src_sc, base_body))
    if src_sc:
        print("src_sc: ", src_sc)
        for body in pm.ls(sl=1):
            print("working on body: ", body)
            shapes = pm.listRelatives(body, c=1)
            print("shapes: ", shapes)
            if shapes:
                shape = shapes[0]
                print("shape: ", shape)
                dst_sc = pm.listConnections(shape, type="skinCluster")
                if dst_sc:
                    dst_sc = dst_sc[0]
                    pm.copySkinWeights(ss=src_sc, ds=dst_sc, sa='closestPoint', noMirror=True)
 
class crowd_group(object):
    def __init__(self, name, mesh_list=None):
        self.name = name
        self.mesh_list = mesh_list or []
    def __repr__(self):
        return '<crowd_group name: %s, meshes: %s>' % (self.name, self.mesh_list)
    def __str__(self):
        output = '  GROUP "%s"\n' % self.name
        if self.mesh_list:
            for word in self.mesh_list:
                output += "   %s\n" % word
        return output

class crowd_model(object):
    def __init__(self, name, groups=None):
        self.name = name
        self.groups = groups or []

    def __repr__(self):
        return '<crowd_model name: %s, groups: %s>' % (self.name, self.groups)
 
    def __str__(self):
        output = ' MODEL "%s"\n' % self.name
        if self.groups:
            for word in self.groups:
                output += "%s\n" % word
        return output

class crowd_char(object):
    def __init__(self, name, models=None):
        self.name = name
        self.models = models or []

    def __repr__(self):
        return '<crowd_char name: %s, groups: %s>' % (self.name, self.groups)
 
    def __str__(self):
        output = 'CHAR "%s"\n' % self.name
        if self.models:
            for word in self.models:
                output += "%s\n" % word
        return output

class crowd_combo_gen(object):
    # parse the crowd file and create geo combos based on that
    def __init__(self, *args):
        self.build_UI()

    def build_UI(self):
        TOT_W = 200
        uiName = 'unreal_crowd'
        uiTitle = 'CREATE UNREAL CROWD'
        if pm.window(uiName, exists=1):
            pm.deleteUI(uiName, window=1)
        pm.window(uiName, title=uiTitle, s=1, mb=0, mxb=0, mnb=0, w=TOT_W, h=200)
        self.CB = {}
        with pm.columnLayout():
            with pm.rowColumnLayout():
                self.CB["FEMALE"] = pm.checkBox(l='Female', vis=1, v=1)
                self.CB["MALE"]   = pm.checkBox(l='Male',   vis=1, v=1)
                self.CB["CHILD"]  = pm.checkBox(l='Child',  vis=1, v=1)

            with pm.rowColumnLayout("lyt3", nc=1):
                pm.button('rig_but', label='CREATE UNREAL CROWD RIGS', bgc=auditColor, h=30, w=TOT_W, c=self.create_crowd_rigs)
        pm.showWindow()
    def create_crowd_rigs(self, *args):
        # go thru the data and generate rigs based on the combinations of elements
        print("create_crowd_rigs: ")
        self.get_geos()
        RIG_DICT = {}
        for gender in self.CB.keys():
            if pm.checkBox(self.CB[gender], q=1, v=1):
                print("We need to process %s" % gender)
                for char in self.char_dict[gender]:
                    strip_name = char.name.replace(gender.lower(), "")
                    strip_name = strip_name.replace("skintone", "")
                    strip_name = strip_name.replace("_", "")
                    print("char: ", char.name)
                    for model in char.models:
                        rig_name = gender + "_" + strip_name + "_" + model.name
                        head_list = []
                        body_list = []
                        shoe_list = []
                        combo_list = []
                        for grp in model.groups:
                            if grp.name.lower() == "head":
                                head_list = grp.mesh_list
                            if grp.name.lower() == "body":
                                body_list = grp.mesh_list
                            if grp.name.lower() == "shoes":
                                shoe_list = grp.mesh_list
                        combo_list = zip(head_list, body_list, shoe_list)
                        geo_list = []

                        for sublist in combo_list:
                            for item in sublist:
                                geo_list.append(item)
                        # geo_list = [item for sublist in combo_list for item in sublist]
                        print("%s geo_list: %s" % (rig_name, geo_list))
                        RIG_DICT[rig_name] = geo_list
        for k, v in RIG_DICT.items():
            print("\nRIG[%s] = %s" % (k,v))
            self.create_crowd_rig(k,v)

    def create_crowd_rig(self, rig_name, geo_names):
        print("create_crowd_rig: %s; %s" % (rig_name, geo_names))
        gender = rig_name.split("_")[0].upper()
        rig = CROWD_SRC_RIG[gender]
        # load the rig
        cmds.file(prompt=False)
        pm.newFile(force=True)
        a_file = pm.openFile(rig)
        cmds.file(prompt=True)
        geo_list = []

        for item in geo_names:
            objs = pm.ls(item)
            if objs:
                # have to cast as str - otherwise object will not be considered a string
                # and we'll fail to find it in the KEEP dict
                geo_list.append(str(objs[0]))
        print("geo_list: ", geo_list)
        # go thru the KEEP dicts based on gender and add the "value" geo to the list to keep
        # keep_dict = KEEP_GEO[gender]
        keep_dict = FEMALE_KEEP_GEO
        keep_geo_list = []

        for geo in geo_list:
            if geo in keep_dict.keys():
                keep_geo_list.extend(keep_dict[geo])
 
        if keep_geo_list:
            geo_list.extend(keep_geo_list)

        # pick a hair geo
        hair_list = pm.ls("hair_*", type="mesh", r=1)
        print("hair_list: ", hair_list)
        hair_geo_list = []

        for item in hair_list:
            hair_geo_list.append(self.get_trans_from_mesh(item))
        print("hair_geo_list: ", hair_geo_list)
        num_items = len(hair_geo_list)
        hair_idx = rand.randint(0, num_items)
        print("keep hair: ", hair_geo_list[hair_idx])
        geo_list.append(hair_geo_list[hair_idx])                
        print("geo_list now: ", geo_list)

        # del all meshes that aren't in geo_list
        all_meshes = pm.ls(type="mesh")
        all_geo = []

        for item in all_meshes:
            all_geo.append(self.get_trans_from_mesh(item))
        del_geo = []

        for item in all_geo:
            if not item in geo_list and not item in del_geo:
                del_geo.append(item)
        print("we should delete: ", del_geo)
        pm.delete(del_geo)
        rig_utils.delete_unused()
        pm.select(pm.ls("CHARACTER"), hierarchy=1)

        # now save the rig to the Rigs folder
        fbx_name = rig_name.replace("_skintone", "")
        fbx_name = "crowd_" + fbx_name
        export_file = os.path.join(EXPORT_RIG_PATH, fbx_name).replace("\\","/")
        print("export_file: ", export_file)
        export_file = export_file + ".fbx"
        print("Exporting: %s" % export_file)

        # check out the file if it exists
        '''
        p4c = p4util.getP4obj(export_file, exception_level=1)
        CL_ID = p4util.get_cl_w_desc("Unreal: rig add/update", create=1, p4inst=p4c)['Change']
        try:
            p4util.p4_edit_or_add(export_file, p4inst=p4c, changeid=CL_ID)
        except P4Exception:
            for e in p4c.errors:
                error = "Error: %s" % e
                print(error)
                return False
        '''
 
        # export fbx file
        # set the properties for exporting the fbx file
        mel_cmd = 'FBXProperty Export|IncludeGrp|CameraGrp|Camera -v false;'  # don't export cameras
        print("mel_cmd = %s" % mel_cmd)
        mel.eval(mel_cmd)
        mel_cmd = 'FBXProperty Export|IncludeGrp|LightGrp|Light -v false;'  # don't export lights
        print("mel_cmd = %s" % mel_cmd)
        mel.eval(mel_cmd)
        mel_cmd = 'FBXProperty Export|IncludeGrp|Animation -v false;'  # make sure animation is OFF
        print("mel_cmd = %s" % mel_cmd)
        mel.eval(mel_cmd)
        mel_cmd = 'FBXProperty Export|AdvOptGrp|UI|ShowWarningsManager -v false;'  # turn off ui warnings
        print("mel_cmd = %s" % mel_cmd)
        mel_cmd = 'FBXProperty Export|AdvOptGrp|Fbx|AsciiFbx -v ASCII;'  # force ascii export
        mel.eval(mel_cmd)
        mel_cmd = 'FBXExport -f "%s" -s;' % export_file  # -s is export selected
        print("mel_cmd = %s" % mel_cmd)
        mel.eval(mel_cmd)

    def get_trans_from_mesh(self, item_name):

        # based on a mesh name, grab the transform above
        par = pm.listRelatives(item_name, p=1)
        if par:
            return str(par[0]) # want the name, not the object
        else:
            return None
 
    def get_geos(self):
        char_list = []
        self.char_dict = {}
        self.reset_vars()
        with open(CROWD_FILE, 'r') as crowd_file:
            for line in crowd_file:
                if line.find("CHARACTER") != -1: # reset everything
                    if self.got_char:
                        char_list.append(crowd_char(char_name, self.model_list))                        
                    self.reset_vars()
                    self.got_char = True
                    char_name = line.split("CHARACTER")[-1]
                    char_name = char_name.replace('"',"")
                    char_name = char_name.split()[0]

                if line.find("MODEL ") != -1: # need space in there to prevent MODEL_SCALE hit
                    if self.got_model: # we were already processing one
                        # need to add last group
                        if self.got_group:
                            self.group_list.append(crowd_group(grp_name, self.mesh_list))
                        self.model_list.append(crowd_model(model_name, self.group_list))
                        self.group_list = []
                        self.mesh_list = []
                        self.got_group = False
                    self.got_model = True
                    model_name = line.split("MODEL")[-1]
                    model_name = model_name.replace('"',"")
                    model_name = model_name.split()[0]

                if line.find("GROUP") != -1:
                    # need to wrap up the current one before processing next one
                    if self.got_group:
                        self.group_list.append(crowd_group(grp_name, self.mesh_list))
                        self.mesh_list = []
                    self.got_group = True
                    grp_name = line.split("GROUP")[-1]
                    grp_name = grp_name.replace('"',"")
                    grp_name = grp_name.split()[0]

                if line.find("MESH_LIST") != -1:
                    if self.got_group:
                        self.mesh_list.extend(self.get_mesh_list(line))

        male_chars = []
        female_chars = []
        child_chars = []

        for char in char_list:

            if char.name.startswith("male"):
                male_chars.append(char)

            if char.name.startswith("female"):
                female_chars.append(char)

            if char.name.startswith("child"):
                child_chars.append(char)

        self.char_dict["MALE"] = male_chars
        self.char_dict["FEMALE"] = female_chars
        self.char_dict["CHILD"] = child_chars

    def get_mesh_list(self, line):
        # turn the mesh list from the input file into separate items
        geos = line.split("=")[-1] # sometimes MESH_LIST is followed by other text so use =
        geos = geos.replace("[","")
        geos = geos.replace("]","")
        geos = geos.replace('"',"")
        geos = geos.strip()
        the_list = [x.strip() for x in geos.split(",")]
        return the_list

    def reset_vars(self):
        self.got_char = False
        self.got_model = False
        self.got_group = False
        self.group_list = []
        self.mesh_list = []
        self.model_list = []
 
class rig_data(object):
    def __init__(self, actor=None, role=None, outfit=None):
        self.actor = actor
        self.role = role
        self.outfit = outfit
        self.cb = None

    def __repr__(self):
        return '<rig_data actor: %s; role: %s; outfit: %s>' % (self.actor, self.role, self.outfit)
 
    def __str__(self):
        output = 'RIG DATA:\n'
        output += '\t' * 2 + 'ACTOR  = %s\n' % self.actor
        output += '\t' * 2 + 'ROLE   = %s\n' % self.role
        output += '\t' * 2 + 'OUTFIT = %s\n' % self.outfit
        return output
 
class create_unreal_rig(object):
    def __init__(self):
        self.p4_CL = None
        self.p4c = None
        self.file_dir = None
        self.ACC_DATA_PATH = anim_utils.get_dev_path(ACC_DATA)
        self.PLYR_DATA_PATH = anim_utils.get_dev_path(PLYR_DATA)
        self.ACC_MAP_PATH = anim_utils.get_dev_path(ACC_MAP)
        self.ACC_DICT = {}

        for line in anim_utils.file_as_list(self.ACC_MAP_PATH):
            pieces = line.split(",")
            if len(pieces) <= 2:
                continue
            key = pieces[1].strip()
            val = pieces[2].strip()
            val = val.replace('"', '') # remove quotes
            self.ACC_DICT[key] = val
        self.create_mocap_rig_ui()

    def create_mocap_rig_ui(self):
        COL_1 = 50
        COL_2 = 350
        TOT_W = COL_1 + COL_2 + 25
        uiName = 'unreal_rig'
        uiTitle = 'CREATE UNREAL RIG'
        if pm.window(uiName, exists=1):
            pm.deleteUI(uiName, window=1)
        pm.window(uiName, title=uiTitle, s=1, mb=0, mxb=0, mnb=0, w=TOT_W, h=200)
        self.cb_list = []
        with pm.frameLayout(cll=0, lv=0, mh=5):
            with pm.rowColumnLayout("lyt1", nc=2, cw=[(1, COL_1), (2, COL_2)]):
                pm.button(l='ID file', c=self.browser)
                self.file_text = pm.textField(text='')
            with pm.rowColumnLayout("lyt3", nc=1):
                pm.button('rig_but', label='CREATE UNREAL RIGS', bgc=auditColor, h=30, w=TOT_W, c=self.create_rigs)
        pm.showWindow()
 
    def browser(self, *args):
        file_list = pm.fileDialog2(fileMode=4, dialogStyle=2, fileFilter="*.*", caption="Text File", okc="Select")

        if file_list:
            print("setting file_dir: ", file_list[0])
            self.file_text.setText(file_list[0])

    def create_rigs(self, *args):
        '''
        updated to FINALLY create both body + face from the same rig
        instead of doing the same stuff twice
        '''

        print("\nCREATE UNREAL RIGS:")
        print("===================")
        the_file = self.file_text.getText().strip()
        if not the_file:
            pm.informBox("Unreal Rig", "No file specified", "OK")
            return False
        file_list = anim_utils.file_as_list(the_file)
        if not file_list:
            pm.informBox("Unreal Rig", "Could not find any files")
            return False
        for graphic_id in file_list:
            start = time.time()
            print("START: ", start)
            self.create_rig(graphic_id)
            end = time.time()
            print("END: ", end)
            secs = (end - start) % 60
            print("%s took %d secs" % (graphic_id, secs))
 
    def create_rig(self, graphic_id):
        '''
        load the sig rig and all of the clothing; then go to sig scale and detach the geo
        then delete everything else so it's just the geo - i.e. no joints, etc.
        then bring in the SOURCE scene so we can bind all the sig geo to that and then
        copy the weighting from the source to the sig geo
        '''
 
        print("\ncreate_rig: ", graphic_id)
        # sync the rig and then load it
        body_rig_path = rig_utils.get_sync_rig_by_id(graphic_id)
        if not body_rig_path:
            print("\t CR: Could not find: ", graphic_id)
            return
        pm.newFile(force=1)
 
        # file should be sync'd, but test it exists just in case
        if not os.path.exists(body_rig_path):
            pm.warning("WARNING: unable to find path: %s" % body_rig_path)
            return
        self.acc_data = self.get_acc_xl_data(self.ACC_DATA_PATH, graphic_id)

        if not self.acc_data:
            pm.informBox("Unreal Rig", "Unable to find accessory info for %s" % graphic_id, "OK")
            return False
        jersey, shorts, team_id = self.get_jersey_shorts(self.PLYR_DATA_PATH, graphic_id)

        if not jersey:
            pm.informBox("Unreal Rig", "Unable to find jersey info for %s" % graphic_id, "OK")
            return False

        if not shorts:
            pm.informBox("Unreal Rig", "Unable to find shorts info for %s" % graphic_id, "OK")
            return False

        if not team_id:
            pm.informBox("Unreal Rig", "Unable to find team id for %s" % graphic_id, "OK")
            return False

        acc_items = self.get_acc_items(self.acc_data)
        acc_items.append(jersey)

        # now try to get the configs.json info so we know what combo of headband/hair/facialhair they should have
        self.json_data = self.get_json_data(graphic_id, body_rig_path)
 
        #+++++++++++++++++++++++++++++++
        # LOAD THE SIG RIG
        #+++++++++++++++++++++++++++++++

        print("\nCR: Loading sig rig: ", body_rig_path)
        new_nodes = cmds.file(body_rig_path, i=1, mergeNamespacesOnClash=0, returnNewNodes=1)
        ref_node = cmds.ls(new_nodes, assemblies=1)[0]
        print("CR: ref_node: ", ref_node)

        # delete headband, hair, facialhair geo that we don't need based on the configs.json info
        self.del_unused_geo(self.json_data)
 
        # need to delete keys on all joints first
        print("==> DELETING KEYS ON ALL JOINTS")
        all_joints = pm.ls(type="joint",r=1)
        pm.cutKey(all_joints)

        #+++++++++++++++++++++++++++++++
        # DRESS UP
        #+++++++++++++++++++++++++++++++

        # import all of the clothing rigs, apply fitting data
        # then detach the geo and delete all of the rigging
        # should be left w/just geo
        # then import the SOURCE scene, bind the sig geo to the source rig
        # then copy the weights from the source to the sig geo and then delete
        # the source geo

        self.occ_faces = self.load_accessories(acc_items)
        print("CR: Apply fitting data to: ", graphic_id)
        bodyCSV.apply_fitting_data(graphic_id)
 
        #########################
        # GET ACTOR INFO
        #########################

        name, gender, skin_color = rig_utils.get_actor_data(graphic_id)
        if not gender:
            print("\t CR: gender was none; resetting to male")
            gender = "male"

        print("\t CR: actor: ", graphic_id)
        print("\t\t CR:  name: ", name)
        print("\t\t CR:  gender: ", gender)
        print("\t\t CR:  skin color: ", skin_color)
 
        #################################
        # IMPORT REFS
        #################################

        # import references so we can delete history on geo and re-parent
        # and delete extra rig shizzle
        print("CR: Importing refs:")
        rig_utils.import_all_refs()
 
        #################################
        # DELETE UNNECESSARY GEO
        #################################

        # first delete all _LOD meshes because we don't care about those
        # and binding/copying weights for all of those takes a lot of extra time
        rig_utils.delete_extra_geo()
 
        #################################
        # DETACHING SIG GEO
        #################################

        print("CR: Unbinding/Detaching Sig Geo:")
        # changed this to unbind since we want to preserve any blendshape data we can
        # so that we might be able to make use of it on the mobu side at some point
        # the head blendshapes are already in use
        # just need to hook up the body and clothing blendshapes

        meshes = rig_utils.get_mesh_list()
        rig_utils.set_scale(scale="unit")
        pm.select(clear=1)
        for mesh in meshes:
            # we need to dupe the baseBody and delete the original because of the shaders and combo shapes
            if "basebody" in mesh.lower():
                print("DUPLICATING: %s" % mesh)
                new_mesh = pm.duplicate(mesh)
                print("new_mesh: ", new_mesh)
                print("deleting ", mesh)
                pm.delete(mesh)
                mesh = new_mesh
                print("mesh now: ", mesh)

            else:  # just delete history because we don't care about blendshape data
                print("CR: Detaching: ", mesh)
                pm.delete(mesh, constructionHistory=1)
        print("CR: Unparenting meshes and deleting sig rig:")

        # unparent the meshes from the sig rig then delete the sig rig
        mesh_grp = rig_utils.find_by_name("meshes")
        if not mesh_grp:
            print("CR: Could not find 'meshes' group!")
            return

        print("CR: mesh_grp = ", mesh_grp)
        pm.parent(mesh_grp, world=1)
 
        # now unparent the meshes so they are all at the root level
        # then delete all the related namespaces
        # and re-group the geo (so they are all on the same level w/o namespaces

        parent_list = []
        meshes = pm.ls(type="mesh")
        for m in meshes:
            dad = pm.listRelatives(m, parent=1)[0]
            if dad not in parent_list:
                parent_list.append(dad)
 
        # now let's delete all unnecessary geo
        new_mesh_list = []
        for p in parent_list:
            if rig_utils.unused_mesh(p):
                print("=> deleting: ", p)
                pm.delete(p)
            else:
                new_mesh_list.append(p)
 
        # move all mesh/transform objects to root level
        pm.parent(new_mesh_list, world=1)

        # delete the old 'meshes' group

        # THERE MIGHT BE MORE THAN 1 MESHES GROUP!
        del_meshes = rig_utils.find_by_name("meshes", find_all=1)
        print("Deleting del_meshes: ", del_meshes)    
        pm.delete(del_meshes)

        # delete any lingering namespaces
        rig_utils.delete_namespaces()
        mesh_grp = pm.group(new_mesh_list, name="dest_meshes")

        ###############################
        # DELETE RIG
        ###############################

        '''
        # since we're now importing the rigged version of the clothes
        # (so we can use the characterPose and fitting data to adjust their size)
        # we have now imported a "player#" for each item of clothing
        # we need to go thru each player node and delete its blendshapeControl node
        # and then delete that rig
        '''
        print("CR: Deleting Rigs:")
 
        # delete the blendshapeControl node first so the meshes

        # don't wind up jacked when we delete the player rig
        objs = pm.ls("blendshapeControl", r=1)
        print("CR: Deleting blendshapeControls: ", objs)
        for obj in objs:
            pm.delete(obj)
 
        # now delete the main rig(s)
        players = pm.ls("player*", assemblies=1)  # only get top-level nodes
        for player in players:
            print("CR: Delete main rig: ", player)
            pm.delete(player)

        # now delete the 'rigging' group and the face stuff under it
        rigging = pm.ls("rigging")
        if rigging:
            pm.delete(rigging)

        print("CR: Delete unused nodes:")
        rig_utils.delete_unused()

        #+++++++++++++++++++++++++++++++++
        # IMPORT SOURCE WEIGHTING SCENE
        #+++++++++++++++++++++++++++++++++

        print("CR: Import source scene: ", gender)
        rig_utils.import_source_scene(gender, use_face=1)
 
        # now set the imported skeleton to the same proportions

        # if we're in Face mode
        try:
            bodyCSV.apply_fitting_data(graphic_id, root_name="hips")
        except:
            print("\t CR: Exception applying fitting data!")
 
        # break connections between face_BSD and the inputs

        # then reset all of the attrs so the face is in the default pose
        if pm.ls('face_BSD'):
            print("CR: Resetting 'face_BSD' node:")
            faceBSD_node = pm.PyNode(pm.ls('face_BSD')[0])
            aliasNames = pm.listAttr('%s.w' % faceBSD_node, multi=1)
            for alias in aliasNames:
                attrName = '%s.%s' % (faceBSD_node, alias)
                pm.disconnectAttr(attrName)
                pm.setAttr(attrName, 0)
        else:
            print("CR: Could not find 'face_BSD' node")

        # delete all BLENDSHAPE NODES!
        blendshapes = pm.ls(type="blendShape")
        print("=> deleting blendshapes: ", blendshapes)
        pm.delete(blendshapes)

        # delete the 'rigging' node which will delete the face rig, etc
        pm.delete(pm.ls("rigging"))
 
        #++++++++++++++++++++++++++++++++++++++
        # TRANSFER WEIGHTS
        #++++++++++++++++++++++++++++++++++++++

        # bind all sig geo to the imported skeleton
        rig_utils.transfer_weights(mesh_grp, graphic_id, use_face=0, shader_copy=False)
 
        # delete extra geo again?
        rig_utils.delete_extra_geo()
 
        # delete any slider joints or dynamic chains
        print("CR: deleting slider joints and dynamic chains:")
        sliders = pm.ls(type="vcSliderJoint",r=1)
        dynamix = pm.ls(type="vcDynamicChain",r=1)
        sliders_grp = pm.listRelatives(sliders, ap=1)
        dynamix_grp = pm.listRelatives(dynamix, ap=1)
        print("deleting: ", sliders_grp)
        pm.delete(sliders_grp)
        print("deleting: ", dynamix_grp)
        pm.delete(dynamix_grp)

        # restore the characterPose - which changes the skeleton to unit scale
        rig_utils.set_scale(scale="unit")
 
        ###############
        # TEXTURES
        ###############

        # run thru all vcFxShader objects and create a lambert for each one?
        convert_vcfx(team_id)

        # now assign the "hidden" shader to the occlusion faces
        self.assign_hidden(self.occ_faces)
 
        # update textures based on skin color
        rig_utils.update_textures(graphic_id, skin_color, name, force_head=False)

        # delete unused nodes again now that we're basically done
        print("CR: Delete unused nodes again:")
        rig_utils.delete_unused()
 
        # and save the result out to the mnRigs folder


        # select everything before exporting
        pm.select(cl=1)
        for node in pm.ls(assemblies=1):
            pm.select(node, add=1)
        export_base_path = body_rig_path.lower().split("sourceart")[0]
        export_path = os.path.join(export_base_path, EXPORT_PATH)
        self.export_rig(export_path, body_rig_path)
        return

    def get_acc_xl_data(self, the_file, graphic_id):

        '''
        get all of the files to swap along w/the frame ranges

        '''

        print("get_acc_xl_data: %s; %s" % (graphic_id, the_file))
        wb = openpyxl.load_workbook(the_file, data_only=1)
        sheet_names = wb.get_sheet_names()
        if not len(sheet_names):
            print("no sheet names found!")
            return
        rig_sheet = "PlayerData"
        if not rig_sheet in sheet_names:
            print("Could not find <%s> sheet in workbook; bailing" % rig_sheet)
            return
 
        ID_COL = 104 # graphic id column
        data_dict = {}
        row_count = 0
        ws = wb[rig_sheet]
        for row in ws.rows:
            if row_count:  # skip first row since it's a header
                cell = row[ID_COL].value or ""
                # print("%s ?= %s" % (cell, graphic_id))
                if str(cell).lower() == "</table>":
                    return data_dict
                if str(cell) != graphic_id:
                    continue
                data_dict[HEADBAND] = row[HEADBAND].value
                data_dict[UNDERSHIRT] = row[UNDERSHIRT].value
                data_dict[SHOES] = row[SHOES].value
                data_dict[SHORTS] = row[SHORTS].value
                data_dict[SOCKS] = row[SOCKS].value
                data_dict[WRIST_L] = row[WRIST_L].value
                data_dict[WRIST_R] = row[WRIST_R].value
                data_dict[ELBOW_L] = row[ELBOW_L].value
                data_dict[ELBOW_R] = row[ELBOW_R].value
                data_dict[ARM_L] = row[ARM_L].value
                data_dict[ARM_R] = row[ARM_R].value
                data_dict[LEG_L] = row[LEG_L].value
                data_dict[LEG_R] = row[LEG_R].value
                data_dict[KNEE_L] = row[KNEE_L].value
                data_dict[KNEE_R] = row[KNEE_R].value
                data_dict[ANKLE_L] = row[ANKLE_L].value
                data_dict[ANKLE_R] = row[ANKLE_R].value
                data_dict[FINGER_L] = row[FINGER_L].value
                data_dict[FINGER_R] = row[FINGER_R].value
                data_dict[ARMTAPE_L] = row[ARMTAPE_L].value
                data_dict[ARMTAPE_R] = row[ARMTAPE_R].value
                data_dict[GOGGLES] = row[GOGGLES].value
                data_dict[FACEMASK] = row[FACEMASK].value
                return data_dict
            row_count += 1
        return

    def get_jersey_shorts(self, the_file, graphic_id):
        print("get_jersey_shorts: %s; %s" % (graphic_id, the_file))

        # grab the TeamData file
        TEAM_DATA_PATH = anim_utils.get_dev_path(TEAM_DATA)
        if not TEAM_DATA_PATH:
            print("Could not find the TeamData file")
            return None, None

        # grab the uniform variables file
        UNI_DATA_PATH = anim_utils.get_src_path(UNI_DIR)
        if not UNI_DATA_PATH:
            print("Could not find the uniform_variables file")
            return None, None

        # now read the PlayerData file
        wb = openpyxl.load_workbook(the_file, data_only=1)
        sheet_names = wb.get_sheet_names()
        if not len(sheet_names):
            print("no sheet names found!")
            return None, None
        rig_sheet = "PlayerData"
        if not rig_sheet in sheet_names:
            print("Could not find <%s> sheet in workbook; bailing" % rig_sheet)
            return None, None

        jersey = None
        REF_COL = 2 # graphic id column
        TEAM_ID = 4
        team_suffix = None
        row_count = 0
        ws = wb[rig_sheet]
        for row in ws.rows:
            if row_count:  # skip first row since it's a header
                cell = row[REF_COL].value or ""
                if str(cell) != graphic_id:
                    continue
                team_suffix = row[TEAM_ID].value
                break
            row_count += 1

        if not team_suffix:
            print("Could not find %s in the PlayerData file" % graphic_id)
            return None, None
        wb = openpyxl.load_workbook(TEAM_DATA_PATH, data_only=1)
        sheet_names = wb.get_sheet_names()

        if not len(sheet_names):
            print("no sheet names found!")
            return None, None
        rig_sheet = "TeamData"

        if not rig_sheet in sheet_names:
            print("Could not find <%s> sheet in workbook; bailing" % rig_sheet)
            return None, None
 
        REF_NAME = 3 # team 2 letter id column
        TEAM_ID = 0 # team # column
        row_count = 0
        ws = wb[rig_sheet]

        for row in ws.rows:
            if row_count:  # skip first row since it's a header
                cell = row[REF_NAME].value or ""

                if str(cell) != team_suffix:
                    continue
                team_prefix = str(row[TEAM_ID].value)
                break
            row_count += 1

        if not team_prefix:
            print("Could not find %s in the TeamData file" % graphic_id)
            return None, None

        if len(team_prefix) == 2:
            team_prefix = "0"+team_prefix
        team_id = team_prefix.lower() + team_suffix.lower()
        print("team_id: ", team_id)

        # now read the uniform_variables file and figure out what jersey/shorts we should be using
        SRC_PATH = 'source path'
        JERSEY_COL = 'jersey type'
        SHORTS_COL = 'custom shorts'
        jersey_type = None
        shorts_type = None
        with open(UNI_DATA_PATH, 'r') as csv_file:
            reader = csv.DictReader(csv_file)
            row_num = 0
            for row in reader:
                src_path = row[SRC_PATH] if row[SRC_PATH] else None

                if not src_path:
                    continue

                if not src_path.lower().startswith(team_id):
                    continue

                if not src_path.lower().endswith("core_home"):
                    continue
                jersey_type = row[JERSEY_COL]
                shorts_type = row[SHORTS_COL]

        # now try to find the jersey in p4
        jersey_name = JERSEY_DICT[jersey_type]
        search_str = '%s/%s/average/%s.ma' % (UNI_PATH, jersey_name, jersey_name)
        jersey_file = anim_utils.find_in_p4(search_str)
        print("jersey_file: ", jersey_file)
 
        # now try to find the shorts in p4
        shorts_name = SHORTS_DICT[shorts_type]
        search_str = '%s/%s/average/%s.ma' % (UNI_PATH, shorts_name, shorts_name)
        shorts_file = anim_utils.find_in_p4(search_str)
        print("shorts_file: ", shorts_file)

        # sync all the relevant data
        files = [constants.CHARACTER_DIR + "/male/clothing/uniforms/textures/%s/current/core_away/core_away_j_color.tga" % team_id,
                    constants.CHARACTER_DIR + "/male/clothing/uniforms/textures/%s/current/core_away/core_away_s_color.tga" % team_id,
                    constants.CHARACTER_DIR + "/male/clothing/uniforms/textures/%s/current/core_home/core_home_j_normal.tga" % team_id,
                    constants.CHARACTER_DIR + "/male/clothing/uniforms/textures/%s/current/core_home/core_home_s_normal.tga" % team_id]
        p4c = p4util.get_p4_obj(p4util.ART_DEPOT, exception_level=1)
        with p4util.p4Connect(p4c):
            try:
                p4c.run_sync(files)
            except:
                pass
        return jersey_file, shorts_file, team_id
 
    def get_acc_items(self, acc_data):
        print("get_acc_items: ")

        # HEADBAND is handled separately
        item_list = []
        for k in acc_data.keys():
            print("%15s: %s" % (KEY_NAMES[k], acc_data[k]))

            # skip SHOES/SOCKS until later
            if k in [SHOES, SOCKS, HEADBAND] or acc_data[k] == "NOITEM" or acc_data[k] == "NONE":
                continue
            if not acc_data[k] in self.ACC_DICT.keys():
                print("==> could not find %s" % KEY_NAMES[k])
                continue
 
            # strip off some items that we don't use
            item_name = self.ACC_DICT[acc_data[k]]
            item_name = item_name.replace("acc_", "_")
            item_name = item_name.replace("_gameday", "")
            if KEY_NAMES[k].startswith("Left"):
                item_name = "l" + item_name
            if KEY_NAMES[k].startswith("Right"):
                item_name = "r" + item_name
            item = self.get_acc_rig(item_name)
            item_list.append(item)

        # do shoes separately
        shoes = self.find_shoes(acc_data[SHOES])
        item_list.append(shoes)

        # do socks separately
        sox = self.find_sox(acc_data[SOCKS])
        item_list.append(sox)
        return item_list

    def load_accessories(self, acc_items):
        # load each accessory, but also get the info from the body_occlusion file
        # so we know which faces on the body should be hidden
        occ_faces = []
        new_faces = rig_utils.import_acc_occ(SHORTS_GEO_PATH, "male")

        if new_faces:
            occ_faces.extend(new_faces)
        for item in acc_items:
            new_faces = rig_utils.import_acc_occ(item, "male")
            if new_faces:
                occ_faces.extend(new_faces)
        return occ_faces

    def find_shoes(self, shoe_code):
        # try to find any file under the shoes dir that contains the shoe_code
        # then go up a few levels and search for the ma file (ma file will not have the code - go figure)
        print("\nfind_shoes: ", shoe_code)
        shoe_path = None
        self.p4c = p4util.get_p4_obj(p4util.ART_DEPOT)
        search_str = '%s/*/average/textures/*/%s_shoe.xml' % (SHOE_DIR, shoe_code)
        shoe_path = anim_utils.find_in_p4(search_str)
        if not shoe_path:
            print("unable to find shoes: %s" % shoe_code)
            return None
        shoe_dir = os.path.dirname(os.path.dirname(os.path.dirname(shoe_path)))
        search_str = '%s/*.ma' % shoe_dir
        return anim_utils.find_in_p4(search_str)

    def find_sox(self, sox_code):
        print("\nfind_sox: ", sox_code)
        sox_path = None
        if not sox_code in SOCKS_DICT.keys():
            pm.warning("Could not find %s in SOCKS dictionary" % sox_code)
            return False
        sox_name = SOCKS_DICT[sox_code]
        self.p4c = p4util.get_p4_obj(p4util.ART_DEPOT)
        search_str = "%s/%s/average/%s.ma" % (ACC_DIR, sox_name, sox_name)
        sox_path = anim_utils.find_in_p4(search_str)
        if not sox_path:
            print("unable to find socks: %s" % sox_code)
            return None
        return sox_path

    def get_json_data(self, graphic_id, body_rig_path):
        print("get_json_data: %s; %s" % (graphic_id, body_rig_path))
        json_path = os.path.dirname(body_rig_path) + "/configs.json"
        json_path = json_path.replace("\\","/")
        json_file = anim_utils.find_in_p4(json_path)
        if not json_file:
            pm.warning("Could not find 'config.json' file for: %s" % graphic_id)
            return None
        json_data = open(json_file).read()
        return json.loads(json_data)

    def assign_hidden(self, occ_faces):
        # find the hidden shader, then assign it to the occ faces
        # sometimes there are multiple hidden shaders so make sure we get the correct one
        print("assign_hidden: ")
        hidden_sg = pm.ls("hidden*", type="shadingEngine")
        if not hidden_sg:
            print("Could not find 'hidden' shader")
            return
        the_sg = None
        for sg in hidden_sg:
            if not "2" in str(sg): # skip shader 2
                the_sg = sg
                break

        # just verify we have baseBody
        body = pm.ls("baseBody")
        if not body:
            pm.warning("Could not find 'baseBody'!")
            return False
        geo_faces = ["baseBody.f[%s]" % f for f in occ_faces]
        print("assigning %s to %s" % (the_sg, geo_faces))
        pm.sets(the_sg, e=1, forceElement=geo_faces)

    def del_unused_geo(self, data): 
        # based on the json data, figure out which
        # headband, hair, facialhair items to keep or delete
        # there will be 2 main keys in the dict: configurations and default_config
        # use the default_config name as a key into the configurations
        if not data:
            return
        configs = data["configurations"]
        default_config = data["default_config"]
        save_geo = []
        for grp in configs:
            if grp["name"] == default_config:
                print("DEFAULT CONFIG: %s" % grp["name"])
                for item in grp["items"]:

                    # there should be items with type of hair, headband, and facialhair
                    if item["type"] == "headband":
                        print("found headband:   ", item["name"])
                        save_geo.append(item["name"])
                    elif item["type"] == "hair":
                        print("found hair:       ", item["name"])
                        save_geo.append(item["name"])
                    elif item["type"] == "facialhair":
                        print("found facialhair: ", item["name"])
                        save_geo.append(item["name"])

        # first get a list of all meshes (shapes) we might want to remove
        all_shapes = []
        all_shapes.extend(pm.ls("headband*", type="mesh", r=1))
        all_shapes.extend(pm.ls("hair*", type="mesh", r=1))
        all_shapes.extend(pm.ls("facial*", type="mesh", r=1))
 
        # then convert to the transform
        all_meshes = []
        for item in all_shapes:
            obj = pm.ls(item, r=1)[0]
            par = pm.listRelatives(obj, parent=1)[0]
            if not par in all_meshes: # don't dupe
                all_meshes.append(par)
 
        # show all of this geo first because what we want to keep might be hidden
        pm.showHidden(all_shapes)
        pm.showHidden(all_meshes)
        del_geo = [x for x in all_meshes if not x in save_geo]
        print("\nto delete: %s" % del_geo)
        for item in del_geo:
            pm.delete(item)

    def get_acc_rig(self, acc_name):
        # search p4 for the specified accessory
        search_str = '%s/*/average/%s.ma' % (ACC_DIR, acc_name)
        return anim_utils.find_in_p4(search_str)

    def export_rig(self, export_path, local_path):
        print("export_rig: ")
        print("\t ", local_path)
 
        # save this to the MOCAP folder
        export_name = os.path.basename(local_path)
        export_name = export_name.split(".")[0]

        # export_name = export_name + ".fbx"
        export_name = export_name + ".usd"
        print("export_name = ", export_name)
        final_path = os.path.join(export_path, export_name)
        final_path = final_path.replace("\\", "/")
        print("final path = ", final_path)
        maya_name = os.path.basename(local_path)
        maya_name = maya_name.split(".")[0]
        maya_name = maya_name + ".ma"
        maya_path = os.path.join(export_path, maya_name)
        maya_path = maya_path.replace("\\", "/")
        print("maya_path = ", maya_path)

        # save a maya file as well:
        pm.saveAs(maya_path, f=1)

        # check to see if this file already exists in P4
        # and we're just updating it or if we need to add it
        p4c = p4util.get_p4_obj(p4util.ART_DEPOT, exception_level=1)
        if os.path.exists(final_path):
            with p4util.p4Connect(p4c):
                CL_ID = p4util.get_cl_w_desc(UPDATE_CL_DESC, create=1, p4inst=p4c)['Change']
                p4c.run_edit("-c", CL_ID, final_path)
        else:
            with p4util.p4Connect(p4c):
                CL_ID = p4util.get_cl_w_desc(CREATE_CL_DESC, create=1, p4inst=p4c)['Change']
                p4c.run_add("-c", CL_ID, final_path)

        # now try to export as USD; grab the CHARACTER node and export
        export_group = asset_utils.get_export_assemblies()
        pm.select(export_group)
        print(export_group)
        pm.mayaUSDExport(f=final_path, selection=1)
        print("Saved ", final_path)
 
        # delete empty changelists
        with p4util.p4Connect(p4c):
            p4c.delete_change(CL_ID)
 
def dupe_meshes():
    # duplicate the eyes, lower_teeth, and tongue geo
    # center the pivots, remove binding, unlock, etc.
    # once in mobu, we're going to parent these to a jaw_joint Null
    # so they can move when the face CONs are adjusted
    # the eyes should move around when _eyeLook_CONs are moved
    # the lower_teeth and tongue should rotate when the jaw opens
    new_r_eye = rig_utils.dupe_obj_by_name("R_eye")
    new_l_eye = rig_utils.dupe_obj_by_name("L_eye")
    new_lower_teeth = rig_utils.dupe_obj_by_name("lower_teeth")
    new_tongue = rig_utils.dupe_obj_by_name("tongue")
 
    # parent the eyes to the head
    head_joint = rig_utils.find_by_name("head")
    pm.parent(new_r_eye, head_joint)
    pm.parent(new_l_eye, head_joint)
 
def convert_vcfx(team_id):
    print("\nconvert_vcfx: ", team_id)
    pm.loadPlugin('vcfxshader', quiet=1)
    pm.select(cl=1)
    vcshaders = pm.ls(type='vcFxShader')
    print("vcshaders: ", vcshaders)
    for vcshader in vcshaders:
        print("\nworking on vcshader: ", vcshader)
        shadinggroups = pm.listConnections('%s.outColor' % vcshader, source=1)
        meshes = []
        for sg in shadinggroups:
            pm.select(sg, replace=1)
            members = pm.ls(sl=1)
            for member in members:
                meshes.append(str(member))
        new_sg = replace_vcfx(vcshader, shadinggroups, team_id)
        print("created new shader: ", new_sg)
        pm.sets(new_sg, forceElement=meshes)
 
def replace_vcfx(vcshader, shadinggroups, team_id):
    print("replace_vcfx: %s; %s; %s" % (vcshader, team_id, shadinggroups))
    name = 'vcfx_' + str(vcshader)
    sn = pm.shadingNode('usdPreviewSurface', asShader=1, name=name)
    for txt in ["ColorTexture", "NormalTexture", "RoughnessTexture", "MaterialColorTexture", "MaterialNormalTexture", "RDMOTexture"]:
        print("RV: working on: %s.%s" % (vcshader, txt))
        if not vcshader.hasAttr(txt):
            # print("  RV: => %s does not have the '%s' attr; skipping" % (vcshader, txt))
            continue
        for con in pm.listConnections("%s.%s" % (vcshader, txt)):
            orig_ftn = pm.getAttr("%s.fileTextureName" % con)
            txt_file_name = vcshader + "_" + txt
            file_name = os.path.basename(orig_ftn).split(".")[0]

            # need to hack the eyes
            if file_name.lower().startswith("iris"):
                orig_ftn = EYES_TXT_DEFAULT
            txt_path = orig_ftn.replace("%databuild_art_root%", DATABUILD_ROOT)
            txt_path = txt_path.replace("\\","/")

            # since the jersey/shorts are generic, we need to override the texture w/the correct team
            if vcshader.lower().startswith("jerseyreg"):
                txt_path = get_jersey_path(team_id)
            if vcshader.lower().startswith("shortsreg"):
                txt_path = get_shorts_path(team_id)

            # make sure we sync the texture!
            get_texture(txt_path)
            print("  RV: txt_path: ", txt_path)
            if txt in ["ColorTexture", "MaterialColorTexture"]:
                node = None
                shader_create_connect_node(sn, node, txt_path, txt_file_name, "outColor", "diffuseColor")
            elif txt in ["NormalTexture", "MaterialNormalTexture"]:
                node = None

                #shader_create_connect_node(sn, node, txt_path, txt_file_name, "outColor", "normal")
                # as a test, try mapping to something other than normal to see if it exports via usd (normal does not currently)
                shader_create_connect_node(sn, node, txt_path, txt_file_name, "outColor", "specularColor")
            elif txt == "RoughnessTexture":
                node = None
                node = shader_create_connect_node(sn, node, txt_path, txt_file_name, "outColor.outColorR", "roughness")

                # node = shader_create_connect_node(sn, node, txt_path, txt_file_name, "outAlpha", "occlusion")
                # node = shader_create_connect_node(sn, node, txt_path, txt_file_name, "outColor.outColorB", "metallic")
            elif txt == "RDMOTexture":
                node = None
                node = shader_create_connect_node(sn, node, txt_path, txt_file_name, "outColor.outColorR", "roughness")
                node = shader_create_connect_node(sn, node, txt_path, txt_file_name, "outAlpha", "occlusion")
                node = shader_create_connect_node(sn, node, txt_path, txt_file_name, "outColor.outColorB", "metallic")
 
    pm.select(cl=1)
    sg = pm.sets(name=name + 'SG', renderable=1)
    pm.disconnectAttr(name + 'SG.surfaceShader')
    try:
        pm.connectAttr(name + '.outColor', name + 'SG.surfaceShader')
    except:
        print("Problem connecting %s.outColor to %sSG.surfaceShader" % (name, name))

    # delete old shader and shading group and rename
    name = str(vcshader)
    pm.delete(vcshader)
    pm.delete(shadinggroups)
    sn.rename(name)
    sg.rename(name + 'SG')
    return sg

def shader_create_connect_node(shader, fileNode, txt_name, node_name, out_name, in_name):
    print("shader_create_connect_node: %s" % shader)
    if not fileNode:
        fileNode = pm.shadingNode('file', asTexture=1, n=node_name)        
    pm.setAttr('%s.fileTextureName' % fileNode, txt_name, type="string")

    # disconnect just in case there's something already connected
    try:
        pm.disconnectAttr('%s.%s' % (shader, in_name))
    except:
        pass

    # connect
    try:
        print(" => connecting %s.%s to %s.%s" % (fileNode, out_name, shader, in_name))
        pm.connectAttr('%s.%s' % (fileNode, out_name), '%s.%s' % (shader, in_name), f=1)
    except:
        print("  Issue connecting %s to %s?" % (out_name, in_name))
    return fileNode

def get_texture(txt_name):
    p4c = p4util.get_p4_obj(p4util.ART_DEPOT)
    with p4util.p4Connect(p4c):
        try:
            p4c.run_sync(txt_name)
        except:
            pass

def get_jersey_path(team_id):
    # given a team id, return a path to that team's shorts
    # also force the texture to be "away" since it's more colorful
    print("get_jersey_path: %s" % team_id)
    jersey_path = JERSEY_TXT_PATH % team_id
    new_path = jersey_path.lower().replace("core_away", "core_home")
    print("new_path: ", new_path)
    return new_path

def get_shorts_path(team_id):
    # given a team id, return a path to that team's shorts
    print("get_shorts_path: %s" % team_id)
    shorts_path = SHORTS_TXT_PATH % team_id
    new_path = shorts_path.lower().replace("core_away", "core_home")
    print("new_path: ", new_path)
    return new_path